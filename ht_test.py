import dt_algorithm_test
import pathlib
import openpyxl as excel
import datetime
import time
import os
import datetime

import read_datasets

rho_list = [0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09]
theta_list = [0.01, 0.02, 0.05, 0.08, 0.1, 0.15, 0.2, 0.25, 0.30, 0.5]
# rho_list = [0.01, 0.05, 0.09]
# theta_list = [0.05, 0.1, 0.2]


def wright_columns(ws, dataset):
    for i in range(len(rho_list)):
        ws.cell(row=1, column=i + 2).value = rho_list[i]
    for i in range(len(theta_list)):
        ws.cell(row=i + 2, column=1).value = theta_list[i]

    ws.cell(row=12, column=1).value = "theta"
    ws.cell(row=1, column=12).value = "rho"
    ws.cell(row=13, column=2).value = dataset

    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/TEST/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/TEST/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/TEST/" + y_m_d + "/ht_memo")

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()
    return y_m_d


def prepare_output_file(data_name):
    target_name = data_name.split(sep="/")

    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/TEST/{y_m_d}/hyper_turning/{date_time}_ht_test_{target_name[2]}.xlsx"

    return file_name


def main():
    # エクセルシートを用意
    wb_name = prepare_output_file()
    wb = excel.Workbook()
    ws = wb.active
    wright_columns(ws)
    wb.create_sheet(title="train_score")
    ws_train_score = wb["train_score"]
    wright_columns(ws_train_score)

    # --
    for i, rho in enumerate(rho_list):
        for j, theta in enumerate(theta_list):
            st_time = time.time()
            print(f"rho: {rho}, theta: {theta}")

            ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score = dt_algorithm_test.main(rho, theta)
            ws.cell(row=j + 2, column=i + 2).value = ROCAUC_test_score
            ws_train_score.cell(row=j + 2, column=i + 2).value = ROCAUC_train_score

            ed_time = time.time()
            print("time: {:.1f}".format(ed_time - st_time))

    wb.save(wb_name)

    return


if __name__ == "__main__":
    main()
