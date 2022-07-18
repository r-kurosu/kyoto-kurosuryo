import statistics
from numpy.ctypeslib import ndpointer
import numpy as np
import pandas as pd
import pulp
import time, copy, itertools, math, warnings, os, sys
import openpyxl as excel
import datetime
import pathlib
from scipy.spatial import distance
import matplotlib.pyplot as plt
from math import dist
# import cplex
# import gurobipy as gp
import listpModule as lpm

import sys
# sys.path.append("/Applications/CPLEX_Studio221/cplex/python/3.9/x86-64_osx")
# print(sys.path)
# import cplex


CPLEX_PATH = "/Applications/CPLEX_Studio221/cplex/bin/x86-64_osx/cplex"
GUROBI_PATH = "/Users/kurosuryou/gurobi.lic"

# LAMBDA = 10


def read_dataset(data_csv, value_txt):
    ### read files ###
    # read the csv and the observed values
    x = pd.read_csv(data_csv, index_col=0)
    value = pd.read_csv(value_txt)

    ### prepare data set ###
    # prepare CIDs
    CIDs = np.array(x.index)
    # prepare target, train, test arrays
    target = np.array(value['a'])
    # construct dictionary: CID to feature vector
    fv_dict = {}
    for cid, row in zip(CIDs, x.values):
        fv_dict[cid] = row
    # construct dictionary: CID to target value
    target_dict = {}
    for cid, val in zip(np.array(value['CID']), np.array(value['a'])):
        target_dict[cid] = val
    # check CIDs: target_values_filename should contain all CIDs that appear in descriptors_filename
    for cid in CIDs:
        if cid not in target_dict:
            # sys.stderr.write('error: {} misses the target value of CID {}\n'.format(target_values_filename, cid))
            exit(1)

    y = np.array([target_dict[cid] for cid in CIDs])
    return x, y


def pre_problem(x, y, D, K):

    st_time1 = time.time()
    index_A = []
    index_B = []

    for i in range(D):
        if y.loc[i] == 0:
            index_A.append(i)
        elif y.loc[i] == 1:
            index_B.append(i)

    temp_max = 0
    if len(index_A) == 0:
        for j in index_B:
            temp = distance.euclidean(0, x.loc[j])
            if temp_max <= temp:
                temp_max = temp
                x_a = [0]*K
                x_b = x.loc[j]
        ed_time1 = time.time()
        print("pre_proc_time = {:.1f}".format(ed_time1 - st_time1))
        return x_a, x_b

    temp_max = 0
    if len(index_B) == 0:
        for i in index_A:
            temp = distance.euclidean(x.loc[i], 0)
            if temp_max <= temp:
                temp_max = temp
                x_a = x.loc[i]
                x_b = [0]*K
        ed_time1 = time.time()
        print("pre_proc_time = {:.1f}".format(ed_time1 - st_time1))
        return x_a, x_b

    temp = 0
    temp_max = 0
    for i in index_A:
        for j in index_B:
            temp = distance.euclidean(x.loc[i], x.loc[j])
            # temp = distance.sqeuclidean(x.loc[i], x.loc[j], 1)
            # temp = np.linalg.norm(x.loc[i]-x.loc[j])
            # temp = np.sqrt(np.power(x.loc[i] - x.loc[j], 2).sum())

            # time1 = time.time()
            # x_i = x.loc[i].values.tolist()
            # x_j = x.loc[j].values.tolist()
            # x_list = x_i + x_j
            # temp = lpm.sum_list(x_list)
            # print("pre_proc_time = {:.1f}".format(time.time() - time1))

            if temp_max <= temp:
                # temp_max = temp
                temp_max = copy.deepcopy(temp)
                x_a = x.loc[i]
                x_b = x.loc[j]
    # print(x_a, x_b)

    ed_time1 = time.time()
    print("pre_proc_time = {:.1f}".format(ed_time1-st_time1))

    return x_a, x_b


def find_separator(x_df, y, D, K, w_p, b_p, x_a, x_b):
    model = pulp.LpProblem("Linear_Separator", pulp.LpMinimize)
    # C = 0.0001

    # 変数定義
    b = pulp.LpVariable("b", cat=pulp.LpContinuous)
    w = [pulp.LpVariable("w_{}".format(i), cat=pulp.LpContinuous) for i in range(K)]
    eps = pulp.LpVariable('eps', lowBound=0, cat=pulp.LpContinuous)
    # eps2 = [pulp.LpVariable("eps_{}".format(i), lowBound=0, cat=pulp.LpContinuous) for i in range(D)]

    # 目的関数
    model += eps
    # model += eps + C*pulp.lpSum(eps2)

    # 制約条件
    model += pulp.lpDot(w, x_a) - b <= -1
    model += pulp.lpDot(w, x_b) - b >= 1
    for i in range(D):
        if y.loc[i] == 0:
            model += pulp.lpDot(w, x_df.loc[i]) - b <= -1 + eps
            # model += pulp.lpDot(w, x_df.loc[i]) - b <= -1 + eps2[i]
        else:
            model += pulp.lpDot(w, x_df.loc[i]) - b >= 1 - eps
            # model += pulp.lpDot(w, x_df.loc[i]) - b >= 1 - eps2[i]

    # model += pulp.lpDot(w, x_a) - b <= -1
    # model += pulp.lpDot(w, x_b) - b >= 1

    status = model.solve(pulp.CPLEX_CMD(path=CPLEX_PATH, msg=0))
    # status = model.solve(pulp.GUROBI(path=GUROBI_PATH, msg=0))

    # 出力
    if status == pulp.LpStatusOptimal:
        w_ast = [w[i].value() for i in range(len(w))]
        b_ast = b.value()
        eps_ast = eps.value()
        w_p.append(w_ast)
        b_p.append(b_ast)
        for i in range(len(w_ast)):
            if w_ast[i] is None:
                w_ast[i] = 0

        return w_ast, b_ast, eps_ast
    else:
        print('FindSeparator infeasible')
        return None, None, None


def new_find_separator(x_df, y, D, K, w_p, b_p, c_arg):
    model = pulp.LpProblem("Linear_Separator", pulp.LpMinimize)

    # 変数定義
    b = pulp.LpVariable("b", cat=pulp.LpContinuous)
    w = [pulp.LpVariable("w_{}".format(i), lowBound=0, cat=pulp.LpContinuous) for i in range(K)]
    w_ = [pulp.LpVariable("w__{}".format(i), lowBound=0, cat=pulp.LpContinuous) for i in range(K)]
    eps = [pulp.LpVariable("eps_{}".format(i), lowBound=0, cat=pulp.LpContinuous) for i in range(D)]

    # 目的関数
    model += pulp.lpSum(w) + pulp.lpSum(w_) + c_arg*(pulp.lpSum(eps))

    # 制約条件
    for i in range(D):
        if y.loc[i] == 0:
            model += pulp.lpDot(w, x_df.loc[i]) - pulp.lpDot(w_, x_df.loc[i]) - b <= -1 + eps[i]
        else:
            model += pulp.lpDot(w, x_df.loc[i]) - pulp.lpDot(w_, x_df.loc[i]) - b >= 1 - eps[i]

    status = model.solve(pulp.CPLEX_CMD(path=CPLEX_PATH, msg=0))
    # status = model.solve(pulp.GUROBI(path=GUROBI_PATH, msg=0))

    # 出力
    if status == pulp.LpStatusOptimal:
        w_ast = [w[i].value() - w_[i].value() for i in range(len(w))]
        b_ast = b.value()
        eps_ast = [eps[i].value() for i in range(D)]
        w_p.append(w_ast)
        b_p.append(b_ast)
        for i in range(len(w_ast)):
            if w_ast[i] is None:
                w_ast[i] = 0

        return w_ast, b_ast, eps_ast
    else:
        print('FindSeparator infeasible')
        return None, None, None

# def find_separator_by_CPLEX(x_df, y, D, K, w_p, b_p, CIDs):
#     model = cplex.Cplex()
#     model.set_problem_type(model.problem_type.MILP)
#     model.objective.set_sense(model.objective.sense.minimize)
#     model.set_problem_name("MILP")
#
#     # 変数定義
#     model.variables.add(names="b", types = "C")
#     [model.variables.add(names=f"w_{i}", types = "C") for i in range(K)]
#     model.variables.add(names="eps", lb=0, types="C")
#
#     # 目的関数
#     model.objective.set_linear("eps", 1)
#
#     # 制約条件
#
#     return


# def find_separator_by_GUROBI(x_df, y, D, K, w_p, b_p, CIDs):
#     model = gp.Model(name = "MILP")
    
    # ノーマルバージョン -----------------------------------------------------------------------
    # w = [0]*K
    # for i in range(K):
    #     # w[i] = model.addVar(vtype=gp.GRB.CONTINUOUS)
    #     w[i] = model.addVar(lb=-1, ub=1, vtype=gp.GRB.CONTINUOUS)
    # # b = model.addVar(vtype=gp.GRB.CONTINUOUS)
    # b = model.addVar(lb=-1, ub=1, vtype=gp.GRB.CONTINUOUS)
    # eps = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)

    # c0 = []
    # c1 = []
    # for i in range(D):
    #     if y.loc[i] == 0:
    #         c0.append(model.addConstr(naiseki(w, x_df.loc[i]) - b <= -1 + eps))
    #     elif y.loc[i] == 1:
    #         c1.append(model.addConstr(naiseki(w, x_df.loc[i]) -b >= 1 - eps))
    # c = model.addConstr(gp.quicksum(w[i] for i in range(K)) > 0)
    # ----------------------------------------------------------------------------------------
    
    
    # 絶対値バージョン-------------------------------------------------------------------------
    # w = [0]*K
    # w_ = [0]*K
    # for i in range(K):
    #     w[i] = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)
    #     w_[i] = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)
    # b = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)
    # b_ = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)
    # eps = model.addVar(lb=0, vtype=gp.GRB.CONTINUOUS)
    # c0 = []
    # c1 = []
    # for i in range(D):
    #     if y.loc[i] == 0:
    #         c0.append(model.addConstr(naiseki(w, x_df.loc[i]) - naiseki(w_, x_df.loc[i]) - b <= -1 + eps))
    #     elif y.loc[i] == 1:
    #         c1.append(model.addConstr(naiseki(w, x_df.loc[i]) - naiseki(w_, x_df.loc[i]) - b >= 1 - eps))
    # c = model.addConstr(gp.quicksum(w[i] for i in range(K)) + gp.quicksum(w_[i] for i in range(K)) >= 0.1)
    # ----------------------------------------------------------------------------------------


    # # 新しい定式化---
    # w = [0]*K
    # for i in range(K):
    #     # w[i] = model.addVar(vtype=gp.GRB.CONTINUOUS)
    #     w[i] = model.addVar(vtype=gp.GRB.CONTINUOUS)
    # # b = model.addVar(vtype=gp.GRB.CONTINUOUS)
    # b = model.addVar(vtype=gp.GRB.CONTINUOUS)
    #
    # # 目的関数
    # # model.setObjective(eps, sense=gp.GRB.MINIMIZE)
    # model.setObjective(sum((naiseki(w, x_df.loc[i])-b - y.loc[i])**2 for i in range(D)))
    #
    # # 求解
    # model.update
    # model.params.NonConvex = 2
    # model.Params.NumericFocus = 3
    # model.Params.OutputFlag == 0
    # model.Params.MIPFocus = 3
    # model.optimize()
    #
    # if model.Status == gp.GRB.OPTIMAL:
    #     w_ast = [w[i].X for i in range(len(w))]
    #     b_ast = b.X
    #     # eps_ast = eps.X
    #     eps_ast = 1
    #     w_p.append(w_ast)
    #     b_p.append(b_ast)
    #     for i in range(len(w_ast)):
    #         if w_ast[i] is None:
    #             w_ast[i] = 0
    #
    #     return w_ast, b_ast, eps_ast
    # else:
    #     print('FindSeparator infeasible')
    #     return None, None, None
    # return


# def find_separator_b(x, y, D, K, w_p, b_p, CIDs):
#     y_obs = 1/2
#     # x_df, y, D, K, w_p, b_p, CIDs
#     MILP = pulp.LpProblem("find_hyperplane", pulp.LpMinimize)
#
#
#     # w = {i: pulp.LpVariable(f"w({i})", cat=pulp.LpContinuous) for i in range(K)}
#     # w = {i: pulp.LpVariable(f"w({i})", -1, 1, cat=pulp.LpContinuous) for i in range(K)}
#     w = [pulp.LpVariable(f"w_{i}", cat=pulp.LpContinuous) for i in range(K)]
#     b = pulp.LpVariable("b", cat=pulp.LpContinuous)
#     eps = pulp.LpVariable("epsilon", 0, cat=pulp.LpContinuous)
#
#     MILP += eps
#
#     # for (_x, _y) in zip(x, y):
#     for i in range(D):
#         _x = x.loc[i]
#         _y = y.loc[i]
#         if _y <= y_obs:
#             # MILP += pulp.lpSum(w[i] * _x[i] for i in range(K)) - b <= 1/4 + eps
#             MILP += pulp.lpDot(w, _x) - b <= -1 + eps
#         else:
#             # MILP += pulp.lpSum(w[i] * _x[i] for i in range(K)) - b >= -0.25 - eps
#             MILP += pulp.lpDot(w, _x) - b >= 1 - eps
#
#     # MILP.writeLP("test.lp")
#
#     # Solve MILP
#     solve_begin = time.time()
#     CPLEX = pulp.CPLEX(path = CPLEX_PATH, msg = 0)
#         # else:
#         #     CPLEX = pulp.CPLEX(path = CPLEX_PATH,
#         #                        msg = 0)
#         # print("Start Solving Using CPLEX...")
#     status = MILP.solve(CPLEX)
#     solve_end = time.time()
#     # else:
#     #     # print("Start Solving Using Coin-OR...")
#     #     solver = pulp.COIN_CMD(msg = CPLEX_MSG)
#     #     MILP.solve(solver)
#     #     solve_end = time.time()
#     if status == pulp.LpStatusOptimal:
#
#         w_value = {i: w[i].value() if w[i].value() is not None else 0 for i in range(K)}
#         b_value = b.value()
#
#         eps_value = eps.value()
#
#     return w_value, b_value, eps_value


def count_s(y):
    s, s_p = 0, 0
    for a in y:
        if a == 0:
            s += 1
        else:
            s_p += 1

    return s, s_p


def sort_dataset(x_df, y, w, b):
    z, z_p = [], []
    for index, a in y.iteritems():
        if a == 0:
            z.append(naiseki(w, x_df.loc[index]) - b)
        elif a == 1:
            z_p.append(naiseki(w, x_df.loc[index]) - b)
    if y is None:
        return z, z_p
    z.sort()
    z_p.sort(reverse=True)

    return z, z_p


def naiseki(a, b):
    result = 0
    for i in range(len(a)):
        result += a[i] * b[i]

    return result


def find_r(z, z_p):
    r, r_p = -1, -1

    for i in range(len(z)):
        if z[i] < 0:
            r = i
        else:
            # print('all z is grater than 0')
            break
    for i in range(len(z_p)):
        if z_p[i] > 0:
            r_p = i
        else:
            # print('all z is less than 0')
            break

    return r, r_p


def find_index_l(z, r, z_p, r_p, s, s_p, rho_arg, theta_arg):
    rho = rho_arg
    theta = theta_arg

    if r == -1:
        c_A = 0
    if r_p == -1:
        c_B = 0

    # if r == -1 or r_p == -1:
        # print(f"can't find l (r=-1 or r'=-1)")
        # print(z)
        # print(z_p)
        # sys.exit()

    # print(f"|z| = {len(z)}, r*theta = {math.floor(r*theta)}")
    # print(f"|z'| = {len(z_p)}, r*theta = {math.floor(r_p*theta)}")

    if r != -1:
        c_A = -z[math.floor(r * theta)]
        for l in range(r+1, 0, -1):
            if siki_1(l, z, z_p, r, r_p, s, s_p, rho) == True:
                # print("find suitable c_A")
                c_A = -z[l-1]
                break
            # if l == 1:
                # print(f"can't find l")

    if r_p != -1:
        c_B = z_p[math.floor(r_p * theta)]
        for l in range(r_p+1, 0, -1):
            if siki_2(l, z, z_p, r, r_p, s, s_p, rho) == True:
                # print("find suitable c_B")
                c_B = z_p[l-1]
                break
            # if l == 1:
                # print(f"can't find l' ")

    return c_A, c_B


def siki_1(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0  # |l|
    # print(r_p, s_p)
    # print(len(z))
    # print(len(z_p))
    for j in range(r_p + 1, s_p):
        if z_p[j] <= z[l-1]:
            temp += 1
    if temp / l <= rho * s_p / s:
        return True

    return False


def siki_2(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0
    for j in range(r + 1, s):
        if z[j] >= z_p[l-1]:
            temp += 1
    if temp / l <= rho * s / s_p:
        return True

    return False


def check_mono(y):
    if (y == 1).sum() == 0:
        print("all 0")
        return True
    if (y == 0).sum() == 0:
        print("all 1")
        return True

    return False


def redefine_func(x, y, w, b, c_A, c_B, CIDs, a_score, f_score, lambda_arg, test_flag):
    # z, z_p = sort_dataset(x, y, w, b)

    wx_b = [0]*len(x)
    for i in range(len(x)):
        wx_b[i] = naiseki(w, x.loc[i]) - b
    z_array = np.array(wx_b)

    order = np.argsort(z_array)
    order_list = order.tolist()

    new_index_0 = []
    new_index_1 = []
    for i in order_list:
        if wx_b[i] > -c_A:
            break
        new_index_0.append(i)
        # print(y[i])
    for i in reversed(order_list):
        if wx_b[i] < c_B:
            break
        new_index_1.append(i)

    if test_flag == 0 or 1:
        if len(new_index_0) != 0:
            pure_0 = pure_rate(y, new_index_0, 0)
        if len(new_index_1) != 0:
            pure_1 = pure_rate(y, new_index_1, 1)
        # z, z_p = sort_dataset(x, y, w, b)
        # if pure_0 < 100 or pure_1 < 100:
        #     print(-c_A, c_B)
        #     print(f"w = {w}")
        #     print(f" z = {z}")
        #     print(f"z' = {z_p}")
        #     plot_func(z, z_p, c_A, c_B)

    for i in new_index_0:
        a_score[CIDs.loc[i]] = 0
        f_score[CIDs.loc[i]] = (1 - pure_0) / 100

    for i in new_index_1:
        a_score[CIDs.loc[i]] = 1
        f_score[CIDs.loc[i]] = pure_1 / 100

    new_index_0.extend(new_index_1)
    drop_index_list = new_index_0

    x.drop(drop_index_list, axis=0, inplace=True)
    y.drop(drop_index_list, inplace=True)
    CIDs.drop(drop_index_list, inplace=True)

    D = len(y)

    return D, x, y, a_score, f_score


def pure_rate(y, index_list, a):
    temp = 0
    for i in index_list:
        if y.loc[i] == a:
            temp += 1
    pure_ = temp / len(index_list) * 100
    # print(f"the rate of class {a} : {pure_}%")

    return pure_


def decision_a_q(y):
    countB = y.sum()
    countA = len(y) - countB
    if countA >= countB:
        a_q = 0
    else:
        a_q = 1

    return a_q


def set_a_q(a_q, a_score):
    a_score.replace(-1, a_q, inplace=True)

    return a_score


def set_a_q_for_f(y, f_score):
    countB = y.sum()
    countA = len(y) - countB

    countA = round(countA, 5)
    countB = round(countB, 5)

    rate = countB / (countA + countB)

    f_score.replace(-1, rate, inplace=True)

    return f_score


def plot_func(z, z_p, c_A, c_B):
    x_0 = [0]*len(z)
    x_1 = [1]*len(z_p)

    plt.scatter(x_0, z)
    plt.scatter(x_1, z_p, c="red")
    plt.xlim([-1, 2])
    plt.hlines(0, -1, 2, "black")
    plt.hlines(-c_A, -1, 2, "b", linestyles="dashed")
    plt.hlines(c_B, -1, 2, "r", linestyles="dashed")

    plt.show()
    # plt.savefig("image.png")

    # sys.exit()

    return


def plot_roc_curv(y_true_test, f_score_test):
    from sklearn.metrics import roc_curve

    # roc = roc_curve(y_true_test, f_score_test)
    fpr, tpr, thresholds = roc_curve(y_true_test, f_score_test)
    print(fpr, tpr, thresholds)
    plt.plot(fpr, tpr, marker='o')
    plt.xlabel('FPR: False positive rate')
    plt.ylabel('TPR: True positive rate')
    plt.grid()
    plt.savefig('sklearn_roc_curve.png')

    return

def experiment_test(x_test, y_test, w, b, c_A, c_B, CIDs_test, a_score_test, f_score_test, rho_arg, theta_arg, lambda_arg):

    # 3.1 s, s'を数える
    # s, s_p = count_s(y_test)
    # 3.2 ソートする
    # z, z_p = sort_dataset(x_test, y_test, w, b)
    # 3.3 r, r'を探す
    # r, r_p = find_r(z, z_p)
    # 3.4 index(l)を探す
    # c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p, rho_arg, theta_arg)

    # 3.5 振り分け
    D, x_test, y_test, a_score_test, f_score_test = redefine_func(x_test, y_test, w, b, c_A, c_B, CIDs_test, a_score_test, f_score_test, lambda_arg, 1)
    # print(a_score_test)
    new_D = len(x_test)

    return new_D, x_test, y_test


def constructing_DT_based_HP(x_df, y, D, K, w_p, b_p, c_p_A, c_p_B, CIDs, a_score, f_score, rho_arg, theta_arg, lambda_arg, c_arg):
    # 2. 超平面（hyper plane）を探す
    x_a, x_b = pre_problem(x_df, y, D, K)
    w, b, eps = find_separator(x_df, y, D, K, w_p, b_p, x_a, x_b)
    print(f"eps = {eps}")
    # print(f"w = {w}")
    # print(f"b = {b}")
    # w, b, eps = new_find_separator(x_df, y, D, K, w_p, b_p, c_arg)

    # print(f"w={w}")
    # print(f"b={b}")
    # print(f"eps={eps}")

    # 3. 決定木の実装
    # 3.1 s, s'を数える
    s, s_p = count_s(y)

    # 3.2 ソートする
    z, z_p = sort_dataset(x_df, y, w, b)

    # 3.3 r, r'を探す
    r, r_p = find_r(z, z_p)

    # 3.4 index(l)を探す
    c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p, rho_arg, theta_arg)
    c_p_A.append(c_A)
    c_p_B.append(c_B)
    # print(c_A, c_B)

    # 3.5 データのプロット
    # plot_func(z, z_p, c_A, c_B)

    # if -c_A == c_B:
        # print(f"c_A = c_B = {-c_A}!!!===================================")

    # if r == -1 or r_p == -1:
    #     print(f"w = {w}")
    #     print(f"b = {b}")


    # 3.5 関数Φとデータセットの再定義
    D, x_df, y, a_score, f_score = redefine_func(x_df, y, w, b, c_A, c_B, CIDs, a_score, f_score, lambda_arg, 0)

    return D, x_df, y


def output_xlx(ws, i, data_name, ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score, depth):
    ws["A"+str(i+2)] = data_name
    ws["B"+str(i+2)] = ROCAUC_train_score
    ws["C"+str(i+2)] = ROCAUC_test_score
    ws["D"+str(i+2)] = BACC_train_score
    ws["E"+str(i+2)] = BACC_test_score
    ws["F"+str(i+2)] = depth

    return


def wright_columns(ws):
    ws["B1"] = "train score ROC/AOC"
    ws["C1"] = "test score ROC/AOC"
    ws["D1"] = "train score BACC"
    ws["E1"] = "test score BACC"
    ws["F1"] = "max depth"

    return


def wright_parameter(ws, rho_arg, theta_arg, n_arg):
    ws["G2"] = f"rho = {rho_arg}"
    ws["G3"] = f"theta = {theta_arg}"
    ws["G4"] = f"n_least = {n_arg}"

    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/CV/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/CV/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/CV/" + y_m_d + "/ht_memo")
    p_file_for_test = pathlib.Path("outputfile/TEST/" + y_m_d)

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()
    if not p_file_for_test.exists():
        p_file_for_test.mkdir()
    return y_m_d


def make_dir_for_test(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/TEST/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/TEST/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/TEST/" + y_m_d + "/ht_memo")

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()

    return y_m_d


def prepare_output_file_for_ht_memo():
    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/CV/{y_m_d}/ht_memo/{date_time}.xlsx"

    return file_name


def prepare_output_file_for_sum():
    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/CV/{y_m_d}/{date_time}_sum.xlsx"

    return file_name


def prepare_output_file_for_test():
    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/TEST/{y_m_d}/{date_time}.xlsx"

    return file_name


def check_exist_dataset_for_test(fail1, fail2, fail3, fail4):
    if os.path.exists(fail1) == False or os.path.exists(fail2) == False or os.path.exists(
            fail3) == False or os.path.exists(fail4) == False:
        return False

    return True


def check_exist_dataset_for_cv(fail1, fail2):
    if os.path.exists(fail1) == False or os.path.exists(fail2) == False:
        return False

    return True
