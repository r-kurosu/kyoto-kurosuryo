import time

import numpy as np

import read_datasets, dt_for_cv, xgb_cv
import pathlib
import openpyxl as excel
import datetime
import time
import random
import os
import sys
import io


depth_list = [3, 4, 5, 6, 7, 8]
eta_list = [0.01, 0.05, 0.1, 0.2, 0.3, 0.5]
boost_list = [500]
weight_list = [0.1, 0.5, 1]
subsample_list = [0.2, 0.5, 0.7, 1]
step_list = [0, 1, 5, 10, 50]

# depth_list = [0.05, 1]
# eta_list = [0.1, 1]
# boost_list = [1, 2]
# depth_list = [0.05]
# eta_list = [0.1]

CV_TIMES = 1


def wright_columns(ws, dataset):
    gap = 10
    ws.cell(row=1, column=2).value = "test"
    ws.cell(row=1, column=2 + gap).value = "train"

    # 1
    ws.cell(row=2, column=1).value = "rho\\theta"
    ws.cell(row=2, column=1 + gap).value = "rho\\theta"
    for i in range(len(eta_list)):
        ws.cell(row=2, column=i + 2).value = eta_list[i]
        ws.cell(row=2, column=i + 2 + gap).value = eta_list[i]
    for i in range(len(depth_list)):
        ws.cell(row=i + 3, column=1).value = depth_list[i]
        ws.cell(row=i + 3, column=1 + gap).value = depth_list[i]
    ws.cell(row=3 + len(eta_list), column=2).value = "(lambda=1)"

    # 2
    ws.cell(row=2 + gap, column=1).value = "rh\\lambda"
    ws.cell(row=2 + gap, column=1 + gap).value = "rho\\lambda"
    for i in range(len(boost_list)):
        ws.cell(row=2 + gap, column=i + 2).value = boost_list[i]
        ws.cell(row=2 + gap, column=i + 2 + gap).value = boost_list[i]
    for i in range(len(depth_list)):
        ws.cell(row=i + 3 + gap, column=1).value = depth_list[i]
        ws.cell(row=i + 3 + gap, column=1 + gap).value = depth_list[i]
    ws.cell(row=3 + gap + len(depth_list), column=2).value = "(rho=0.05)"

    # 3
    ws.cell(row=2 + gap*2, column=1).value = "lambda\\theta"
    ws.cell(row=2 + gap*2, column=1 + gap).value = "lambda\\theta"
    for i in range(len(eta_list)):
        ws.cell(row=2 + gap*2, column=i + 2).value = eta_list[i]
        ws.cell(row=2 + gap*2, column=i + 2 + gap).value = eta_list[i]
    for i in range(len(boost_list)):
        ws.cell(row=i + 3 + gap*2, column=1).value = boost_list[i]
        ws.cell(row=i + 3 + gap*2, column=1 + gap).value = boost_list[i]
    ws.cell(row=3 + gap*2 + len(boost_list), column=2).value = "(theta=0.1)"

    # ws.cell(row=12, column=2).value = dataset

    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/XGB/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/XGB/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/XGB/" + y_m_d + "/ht_memo")

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()
    return y_m_d


def prepare_output_file(data_name):
    target_name = data_name.split(sep="/")

    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/XGB/{y_m_d}/hyper_turning/{date_time}_ht_cv_{target_name[2]}.xlsx"

    return file_name


def main():
    INPUT_CSV, INPUT_TXT = read_datasets.read_data_list_for_cv()
    wb_name = [0]*len(INPUT_CSV)
    gap = 10

    for k, data in enumerate(INPUT_CSV):
        ht_start_time = time.time()

        # エクセルシートを用意
        wb_name[k] = prepare_output_file(data_name=data)
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "BACC score"
        wright_columns(ws, data)
        wb.create_sheet(title="AUC score")
        ws2 = wb["AUC score"]
        wright_columns(ws2, data)

        bacc_score_data_test = np.zeros((len(depth_list), len(eta_list), len(boost_list), len(weight_list), len(subsample_list), len(step_list)))
        bacc_score_data_train = np.zeros((len(depth_list), len(eta_list), len(boost_list), len(weight_list), len(subsample_list), len(step_list)))
        auc_score_data_test = np.zeros((len(depth_list), len(eta_list), len(boost_list), len(weight_list), len(subsample_list), len(step_list)))
        auc_score_data_train = np.zeros((len(depth_list), len(eta_list), len(boost_list), len(weight_list), len(subsample_list), len(step_list)))

        # --
        for i, depth in enumerate(depth_list):
            for j, eta in enumerate(eta_list):
                for l, boost in enumerate(boost_list):
                    for m, weight in enumerate(weight_list):
                        for n, subsample in enumerate(subsample_list):
                            for s, step in enumerate(step_list):
                                st_time = time.time()
                                print(f"max_depth:{depth},eta:{eta}, boost:{boost}, weight:{weight}, subsample:{subsample}, step:{step}")

                                ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score = xgb_cv.main(INPUT_CSV[k], INPUT_TXT[k], cv_times=CV_TIMES, max_depth=depth, eta=eta, boost=boost, weight=weight, subsample=subsample, step=step)

                                # 出力
                                if subsample == 1 and step == 0:
                                    if eta == 1:
                                        ws.cell(row=i + 3, column=j + 2).value = BACC_test_score
                                        ws.cell(row=i + 3, column=j + 2 + gap).value = BACC_train_score
                                        ws2.cell(row=i + 3, column=j + 2).value = ROCAUC_test_score
                                        ws2.cell(row=i + 3, column=j + 2 + gap).value = ROCAUC_train_score

                                    if depth == 0.05:
                                        ws.cell(row=j + 3 + gap, column=l + 2).value = BACC_test_score
                                        ws.cell(row=j + 3 + gap, column=l + 2 + gap).value = BACC_train_score
                                        ws2.cell(row=j + 3 + gap, column=l + 2).value = ROCAUC_test_score
                                        ws2.cell(row=j + 3 + gap, column=l + 2 + gap).value = ROCAUC_train_score

                                    if eta == 0.1:
                                        ws.cell(row=l + 3 + gap*2, column=i + 2).value = BACC_test_score
                                        ws.cell(row=l + 3 + gap*2, column=i + 2 + gap).value = BACC_train_score
                                        ws2.cell(row=l + 3 + gap*2, column=i + 2).value = ROCAUC_test_score
                                        ws2.cell(row=l + 3 + gap*2, column=i + 2 + gap).value = ROCAUC_train_score

                                bacc_score_data_test[i, j, l, m, n, s] = BACC_test_score
                                bacc_score_data_train[i, j, l, m, n, s] = BACC_train_score
                                auc_score_data_test[i, j, l, m, n, s] = ROCAUC_test_score
                                auc_score_data_train[i, j, l, m, n, s] = ROCAUC_train_score

                                ed_time = time.time()
                                print("time {:.1f}".format(ed_time - st_time))

        wb.save(wb_name[k])

        bacc_max_test_index = np.unravel_index(np.argmax(bacc_score_data_test), bacc_score_data_test.shape)
        bacc_max_train_index = np.unravel_index(np.argmax(bacc_score_data_train), bacc_score_data_train.shape)
        auc_max_train_index = np.unravel_index(np.argmax(auc_score_data_train), auc_score_data_train.shape)
        auc_max_test_index = np.unravel_index(np.argmax(auc_score_data_test), auc_score_data_test.shape)
        
        # print(max_test_index, max_train_index)
        bacc_max_test_score = bacc_score_data_test[bacc_max_test_index]
        bacc_max_train_score = bacc_score_data_train[bacc_max_train_index]
        auc_max_test_score = auc_score_data_test[auc_max_test_index]
        auc_max_train_score = auc_score_data_train[auc_max_train_index]

        ht_end_time = time.time()

        print("*"*70)
        print(f"{data}")
        print(f"max train score (BACC) = {bacc_max_train_score}, (max_depth={depth_list[bacc_max_train_index[0]]}, eta={eta_list[bacc_max_train_index[1]]}, boost={boost_list[bacc_max_train_index[2]]}, weight={weight_list[bacc_max_train_index[3]]}, subsample={subsample_list[bacc_max_train_index[4]]}, step={step_list[bacc_max_train_index[5]]})")
        print(f"max test score (BACC) = {bacc_max_test_score}, (max_depth={depth_list[bacc_max_test_index[0]]}, eta={eta_list[bacc_max_test_index[1]]}, boost={boost_list[bacc_max_test_index[2]]}, weight={weight_list[bacc_max_test_index[3]]}, subsample={subsample_list[bacc_max_test_index[4]]}, step={step_list[bacc_max_test_index[5]]})")
        print(f"max train score (AUC) = {auc_max_train_score}, (max_depth={depth_list[auc_max_train_index[0]]}, eta={eta_list[auc_max_train_index[1]]}, boost={boost_list[auc_max_train_index[2]]}, weight={weight_list[auc_max_train_index[3]]}, subsample={subsample_list[auc_max_train_index[4]]}, step={step_list[auc_max_train_index[5]]})")
        print(f"max test score (AUC) = {auc_max_test_score}, (max_depth={depth_list[auc_max_test_index[0]]}, eta={eta_list[auc_max_test_index[1]]}, boost={boost_list[auc_max_test_index[2]]}, weight={weight_list[auc_max_test_index[3]]}, subsample={subsample_list[auc_max_test_index[4]]}, step={step_list[auc_max_test_index[5]]})")
        print("計算時間 {:.1f}".format(ht_end_time - ht_start_time))
        print("*"*70)

    return


if __name__ == "__main__":
    main()
