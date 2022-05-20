import statistics

import datetime as datetime
import numpy as np
import pandas as pd
import pulp
import time, copy, itertools, math, warnings, os
import openpyxl as excel
import datetime
import pathlib
from sklearn.linear_model import Lasso, LinearRegression
from sklearn.model_selection import train_test_split, cross_val_score, KFold
from sklearn.metrics import roc_auc_score, balanced_accuracy_score

import dt_tools
import io, sys
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# warnings.simplefilter('ignore')
CPLEX_PATH = "/Applications/CPLEX_Studio221/cplex/bin/x86-64_osx/cplex"

# マクロ定義
RHO = 0.05
THETA = 0.1
N_LEAST = 10
# CSV_DATA = "dataset/AhR_large_var0_quadratic_h5000_desc_norm.csv"
# VALUE_DATA = "dataset/AhR_large_values.txt"
CSV_DATA = "dataset/AhR_large_var0_quadratic_h25000_desc_norm.csv"
VALUE_DATA = "dataset/AhR_large_values.txt"

TIMES = 2 # CVの回数（実験は10で行う）


def read_dataset(data_csv, value_txt):
    ### read files ###
    # read the csv and the observed values
    x = pd.read_csv(data_csv, index_col=0)
    value = pd.read_csv(value_txt)

    ### prepare data set ###
    # prepare CIDs
    CIDs = np.array(x.index)
    # prepare target, train, test arrays
    target = np.array(value['a'])
    # construct dictionary: CID to feature vector
    fv_dict = {}
    for cid, row in zip(CIDs, x.values):
        fv_dict[cid] = row
    # construct dictionary: CID to target value
    target_dict = {}
    for cid, val in zip(np.array(value['CID']), np.array(value['a'])):
        target_dict[cid] = val
    # check CIDs: target_values_filename should contain all CIDs that appear in descriptors_filename
    for cid in CIDs:
        if cid not in target_dict:
            sys.stderr.write('error: {} misses the target value of CID {}\n'.format(target_values_filename, cid))
            exit(1)

    y = np.array([target_dict[cid] for cid in CIDs])
    return x, y


def find_separator(x_df, y, D, K, w_p, b_p, CIDs):
    model = pulp.LpProblem("Linear_Separator", pulp.LpMinimize)
    # 変数定義
    b = pulp.LpVariable("b", -1, 1, cat=pulp.LpContinuous)
    w = [pulp.LpVariable("w_{}".format(i), -1, 1, cat=pulp.LpContinuous) for i in range(K)]
    eps = pulp.LpVariable('eps', cat=pulp.LpContinuous)
    # 目的関数
    model += eps
    # 制約条件
    for i in range(D):
        if y.loc[i] == 0:
            model += pulp.lpDot(w, x_df.loc[i]) - b <= -1 + eps
        else:
            model += pulp.lpDot(w, x_df.loc[i]) - b >= 1 - eps

    status = model.solve(pulp.CPLEX(path=CPLEX_PATH, msg=0))

    # 出力
    if status == pulp.LpStatusOptimal:
        w_ast = [w[i].value() for i in range(len(w))]
        b_ast = b.value()
        eps_ast = eps.value()
        w_p.append(w_ast)
        b_p.append(b_ast)
        for i in range(len(w_ast)):
            if w_ast[i] is None:
                w_ast[i] = 0

        return w_ast, b_ast, eps_ast
    else:
        print('FindSeparator infeasible')
        return None, None, None


def count_s(y):
    s, s_p = 0, 0
    for a in y:
        if a == 0:
            s += 1
        else:
            s_p += 1

    return s, s_p


def sort_dataset(x_df, y, w, b):
    z, z_p = [], []
    # print(x_df)
    # print(y)
    # print(w)
    for index, a in y.iteritems():
        if a == 0:
            z.append(naiseki(w, x_df.loc[index]) - b)
        elif a == 1:
            z_p.append(naiseki(w, x_df.loc[index]) - b)
    if y is None:
        return z, z_p
    z.sort()
    z_p.sort(reverse=True)
    # print(z[:5])
    # print(z_p[:5])

    return z, z_p


def naiseki(a, b):
    result = 0
    for i in range(len(a)):
        result += a[i] * b[i]

    return result


def find_r(z, z_p):
    r, r_p = -1, -1
    for i in range(len(z)):
        if z[i] < 0:
            r = i
        else:
            # print('all z is grater than 0')
            break
    for i in range(len(z_p)):
        if z_p[i] > 0:
            r_p = i
        else:
            # print('all z is smaller than 0')
            break

    return r, r_p


def find_index_l(z, r, z_p, r_p, s, s_p):
    rho = RHO
    theta = THETA
    for l in range(r, 0, -1):
        if siki_1(l, z, z_p, r, r_p, s, s_p, rho) == True:
            # print(f"l={l}")
            # print(f"r={r}")
            c_A = -z[l]
            break
        c_A = -z[math.floor(r*theta)]

    for l in range(r_p, 0, -1):
        if siki_2(l, z, z_p, r, r_p, s, s_p, rho) == True:
            # print(f"l={l}")
            # print(f"r\'={r_p}")
            c_B = -z_p[l]  # TODO: type miss???
            break
        c_B = -z_p[math.floor(r_p*theta)]

    if r <= 0:
        c_A = 0
    if r_p <= 0:
        c_B = 0

    return c_A, c_B


def siki_1(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0 # |l|
    for j in range(r_p + 1, s_p):
        if z_p[j] <= z[l]:
            temp += 1
    if temp / l <= rho * s_p / s:
        return True

    return False


def siki_2(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0
    for j in range(r + 1, s):
        if z[j] >= z_p[l]:
            temp += 1
    if temp / l <= rho * s / s_p:
        return True

    return False


def redefine_func(x_df, y, w, b, c_A, c_B, CIDs, a_score):
    for index, vector_x in x_df.iterrows():
        # ver2. 正解を気にしない
        if naiseki(w, vector_x) - b <= -c_A:
            x_df.drop(index, axis=0, inplace=True)
            y.drop(index, inplace=True)
            a_score[CIDs.loc[index]] = 0
            CIDs.drop(index, inplace=True)

        elif naiseki(w, vector_x) - b >= c_B:
            x_df.drop(index, axis=0, inplace=True)
            y.drop(index, inplace=True)
            a_score[CIDs.loc[index]] = 1
            CIDs.drop(index, inplace=True)

    D = len(y)

    return D, x_df, y


def experiment_test(x_test, y_test, w, b, CIDs_test, a_score_test):
    # 3.1 s, s'を数える
    s, s_p = count_s(y_test)
    # 3.2 ソートする
    z, z_p = sort_dataset(x_test, y_test, w, b)
    # 3.3 r, r'を探す
    r, r_p = find_r(z, z_p)
    # 3.4 index(l)を探す
    c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p)
    # 3.5 振り分け
    for index, vector_x in x_test.iterrows():
        if naiseki(w, vector_x) - b <= -c_A:
            x_test.drop(index, axis=0, inplace=True)
            y_test.drop(index, inplace=True)
            a_score_test[CIDs_test.loc[index]] = 0
            CIDs_test.drop(index, inplace=True)

        elif naiseki(w, vector_x) - b >= c_B:
            x_test.drop(index, axis=0, inplace=True)
            y_test.drop(index, inplace=True)
            a_score_test[CIDs_test.loc[index]] = 1
            CIDs_test.drop(index, inplace=True)

    new_D = len(x_test)
    return new_D, x_test, y_test


def calc_a_q(D, K, a, x):
    a_q = [0]*K
    for d in range(K):
        temp_A = 0
        temp_B = 0
        # for i in range(D):
        #     if a[i] == 0 and x[i][d] isinD_q:
        #         temp_A += 1
        #     elif a[i] == 1 and x[i][d] isinD_q:
        #         temp_B += 1
        if temp_A > temp_B:
            a_q[d] = 0
        elif temp_A < temp_B:
            a_q[d] = 1

    return a_q


def constructing_DT_based_HP(x_df, y, D, K, w_p, b_p, c_p_A, c_p_B, CIDs, a_score):
    # 2. 超平面（hyper plane）を探す
    # w, b = use_sklearn(x, y)
    w, b, eps = find_separator(x_df, y, D, K, w_p, b_p, CIDs)
    # print(f"w={w}")
    # print(f"b={b}")
    # print(f"eps={eps}")

    # 3. 決定木の実装
    # 3.1 s, s'を数える
    s, s_p = count_s(y)

    # 3.2 ソートする
    z, z_p = sort_dataset(x_df, y, w, b)
    # print(s, len(z))
    # print(s_p, len(z_p))

    # 3.3 r, r'を探す
    r, r_p = find_r(z, z_p)
    # print(r, r_p)

    # 3.4 index(l)を探す
    c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p)
    c_p_A.append(c_A)
    c_p_B.append(c_B)
    # print(c_A, c_B)

    # 3.5 関数Φとデータセットの再定義
    D, x_df, y = redefine_func(x_df, y, w, b, c_A, c_B, CIDs, a_score)

    return D, x_df, y


def set_a_q(x_df, y, CIDs, a_score):
    countA, countB = 0, 0
    countB = y.sum()
    countA = len(y) - countB
    if countA >= countB:
        value = 0
    else:
        value = 1

    for index, vector_x in x_df.iterrows():
        a_score[CIDs.loc[index]] = value

    return


def test_main(INPUT_CSV, INPUT_TXT):
    # 1. read data set
    data_csv = INPUT_CSV
    value_text = INPUT_TXT

    x, y = read_dataset(data_csv, value_text)
    x_df = x.reset_index(drop=True)
    y = pd.Series(y)
    CIDs = pd.Series(list(x.index))
    a_score = pd.Series([-1]*len(CIDs), index=list(CIDs))
    # print(x_df)
    # print(x)
    # print(CIDs)

    # (TIMES)回 5-fold回す
    test_scores = []
    train_scores = []
    test_scores_bacc = []
    train_scores_bacc = []
    st_time = time.time()
    for times in range(TIMES):
        # print("-----------------------------------------------")
        # print(f"{times+1}回目の交差実験")
        # 5-foldCVによる分析
        kf = KFold(n_splits=5, shuffle=True, random_state=times)
        # kf = KFold(n_splits=5, shuffle=True, random_state=times+1000)
        ROC_AOC_scores_train = []
        ROC_AOC_scores_test = []
        BACC_scores_train = []
        BACC_scores_test = []

        for train_id, test_id in kf.split(x_df):
            x_train, x_test = x_df.iloc[train_id], x_df.iloc[test_id]
            y_train, y_test = y.iloc[train_id], y.iloc[test_id]
            CIDs_train, CIDs_test = CIDs.iloc[train_id], CIDs.iloc[test_id]
            a_score_train, a_score_test = a_score.iloc[train_id], a_score.iloc[test_id]

            # 2. construct a decision tree using hyper plane
            x_train = x_train.reset_index(drop=True)
            y_true_train = copy.deepcopy(y_train)
            y_train = y_train.reset_index(drop=True)
            CIDs_train.reset_index(drop=True, inplace=True)

            # マクロ変数
            D = len(y_train)  # データの数
            K = len(x_df.columns)  # ベクトルサイズ（記述示の数）
            N_least = N_LEAST  # 許容
            p = 0

            w_p = []
            b_p = []
            c_p_A = []
            c_p_B = []

            while D > N_least:
                p += 1
                # print(f"|D|={D}", f"p={p}")
                new_D, new_x_df, new_y = constructing_DT_based_HP(x_train, y_train, D, K, w_p, b_p, c_p_A, c_p_B, CIDs_train, a_score_train)
                D = new_D
                x_train = new_x_df.reset_index(drop=True)
                y_train = new_y.reset_index(drop=True)
                CIDs_train.reset_index(drop=True, inplace=True)
            q = p+1

            set_a_q(x_train, y_train, CIDs_train, a_score_train)

            # 3. test ---------------------------------------------
            D = len(x_test)
            x_test = x_test.reset_index(drop=True)
            y_true_test = copy.deepcopy(y_test)
            y_test = y_test.reset_index(drop=True)
            CIDs_test.reset_index(drop=True, inplace=True)

            # print(len(b_p))
            for p in range(len(b_p)):
                new_D, new_x_df, new_y = experiment_test(x_test, y_test, w_p[p], b_p[p], CIDs_test, a_score_test)
                D = new_D
                x_test = new_x_df.reset_index(drop=True)
                y_test = new_y.reset_index(drop=True)
                CIDs_test.reset_index(drop=True, inplace=True)

            set_a_q(x_test, y_test, CIDs_test, a_score_test)

            # 4. 結果 -------------------------------
            a_score_train = a_score_train.to_numpy()
            train_score = roc_auc_score(y_true_train, a_score_train)
            bacc_train_score = balanced_accuracy_score(y_true_train, a_score_train)
            train_scores.append(train_score)
            train_scores_bacc.append(bacc_train_score)
            # print(f"ROC/AUC train score: {train_score}")
            # print(f"BACC train score: {bacc_train_score}")
            # ROC_AOC_scores_train.append(train_score)
            # BACC_scores_train.append(train_score)

            a_score_test = a_score_test.to_numpy()
            test_score = roc_auc_score(y_true_test, a_score_test)
            bacc_test_score = balanced_accuracy_score(y_true_test, a_score_test)
            test_scores.append(test_score)
            test_scores_bacc.append(bacc_test_score)
            # print(f"ROC/AUC test score: {test_score}")
            # print(f"BACC test score: {bacc_test_score}")
            # ROC_AOC_scores_test.append(test_score)
            # BACC_scores_test.append(test_score_bacc)
            # -----------------------------------------
        # 5foldCV終了
        # print(f"ROC AUC train score: {statistics.median(ROC_AOC_scores_train)}, {ROC_AOC_scores_train}")
        # print(f"ROC AUC test score: {statistics.median(ROC_AOC_scores_test)}, {ROC_AOC_scores_test}")

    # 10回のCV終了
    ed_time = time.time()
    ROCAUC_train_score = statistics.median(train_scores)
    ROCAUC_test_score = statistics.median(test_scores)
    BACC_train_score = statistics.median(train_scores_bacc)
    BACC_test_score = statistics.median(test_scores_bacc)

    # print("======================================================")
    print(data_csv)
    print(f"ROC/AUC train score (median): {ROCAUC_train_score}")
    print(f"ROC/AUC test score (median): {ROCAUC_test_score}")
    print(f"ROC/AUC train score (median): {BACC_train_score}")
    print(f"ROC/AUC test score (median): {BACC_test_score}")
    print("計算時間 : {:.1f}".format(ed_time - st_time))
    # print("======================================================")

    return ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score


def output_xlx(ws, i, data_name, ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score):
    ws["A"+str(i+2)] = data_name
    ws["B"+str(i+2)] = ROCAUC_train_score
    ws["C"+str(i+2)] = ROCAUC_test_score
    ws["D"+str(i+2)] = BACC_train_score
    ws["E"+str(i+2)] = BACC_test_score

    return


def wright_columns(ws):
    ws["B1"] = "train score ROC/AOC"
    ws["C1"] = "test score ROC/AOC"
    ws["D1"] = "train score BACC"
    ws["E1"] = "test score BACC"

    return


def wright_parameter(ws):
    ws["G2"] = f"rho = {RHO}"
    ws["G3"] = f"theta = {THETA}"
    ws["G4"] = f"n_least = {N_LEAST}"

    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/CV/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/CV/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/CV/" + y_m_d + "/ht_memo")

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()

    return y_m_d


def prepare_output_file():
    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/CV/{y_m_d}/ht_memo/{date_time}.xlsx"

    return file_name


def main(rho_arg, theta_arg, INPUT_CSV, INPUT_TXT):
    global RHO
    global THETA
    RHO, THETA = rho_arg, theta_arg

    # エクセルシートを用意
    wbname1 = prepare_output_file()
    wb = excel.Workbook()
    ws = wb.active
    wright_columns(ws)
    wright_parameter(ws)

    ## 1. experiment all datasets
    # INPUT_CSV, INPUT_TXT = read_data_list()
    ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score = test_main(INPUT_CSV, INPUT_TXT)
    output_xlx(ws, 1, INPUT_CSV, ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score)

    wb.save(wbname1)

    ## 2. experiment one dataset
    # print(f"experiment {CSV_DATA}")
    # test_main(CSV_DATA, VALUE_DATA)
    # wb.save(wbname1)

    return ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score


if __name__ == "__main__":
    INPUT_CSV, INPUT_TXT = read_data_list()
    main(RHO, THETA, INPUT_CSV, INPUT_TXT)
