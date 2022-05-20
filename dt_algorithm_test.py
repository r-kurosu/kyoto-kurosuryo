import statistics

import numpy as np
import pandas as pd
import pulp
import time, copy, itertools, math, warnings, os
import openpyxl as excel
import datetime
import pathlib
import read_datasets

from sklearn.linear_model import Lasso, LinearRegression
from sklearn.model_selection import train_test_split, cross_val_score, KFold
from sklearn.metrics import roc_auc_score, balanced_accuracy_score

import io, sys
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

warnings.simplefilter('ignore')
CPLEX_PATH = "/Applications/CPLEX_Studio221/cplex/bin/x86-64_osx/cplex"

# マクロ定義
RHO = 0.05
THETA = 0.1
N_LEAST = 10
INPUT_DATA = "dataset/classification_var0_5000_42922/AhR_large_var0_quadratic_h200_desc_norm.csv"
VALUE_DATA = "dataset/classification_var0_5000_42922/AhR_large_values.txt"
TEST_INPUT_DATA = "dataset/classification_test_var0_5000_42922/AhR_large_var0_quadratic_h200_desc_norm.csv"
TEST_VALUE_DATA = "dataset/classification_test_var0_5000_42922/AhR_large_values.txt"

# print(f"rho : {RHO}")
# print(f"theta : {THETA}")
# print(f"n_least : {N_LEAST}")


def read_dataset(data_csv, value_txt):
    ### read files ###
    # read the csv and the observed values
    x = pd.read_csv(data_csv, index_col=0)
    value = pd.read_csv(value_txt)

    ### prepare data set ###
    # prepare CIDs
    CIDs = np.array(x.index)
    # prepare target, train, test arrays
    target = np.array(value['a'])
    # construct dictionary: CID to feature vector
    fv_dict = {}
    for cid, row in zip(CIDs, x.values):
        fv_dict[cid] = row
    # construct dictionary: CID to target value
    target_dict = {}
    for cid, val in zip(np.array(value['CID']), np.array(value['a'])):
        target_dict[cid] = val
    # check CIDs: target_values_filename should contain all CIDs that appear in descriptors_filename
    for cid in CIDs:
        if cid not in target_dict:
            sys.stderr.write('error: {} misses the target value of CID {}\n'.format(target_values_filename, cid))
            exit(1)

    y = np.array([target_dict[cid] for cid in CIDs])
    return x, y


def find_separator(x_df, y, D, K, w_p, b_p, CIDs):
    model = pulp.LpProblem("Linear_Separator", pulp.LpMinimize)
    # 変数定義
    b = pulp.LpVariable("b", -1, 1, cat=pulp.LpContinuous)
    w = [pulp.LpVariable("w_{}".format(i), -1, 1, cat=pulp.LpContinuous) for i in range(K)]
    eps = pulp.LpVariable('eps', cat=pulp.LpContinuous)
    # 目的関数
    model += eps
    # 制約条件
    for i in range(D):
        if y.loc[i] == 0:
            model += pulp.lpDot(w, x_df.loc[i]) - b <= -1 + eps
        else:
            model += pulp.lpDot(w, x_df.loc[i]) - b >= 1 - eps

    status = model.solve(pulp.CPLEX(path=CPLEX_PATH, msg=0))

    # 出力
    if status == pulp.LpStatusOptimal:
        w_ast = [w[i].value() for i in range(len(w))]
        b_ast = b.value()
        eps_ast = eps.value()
        w_p.append(w_ast)
        b_p.append(b_ast)
        for i in range(len(w_ast)):
            if w_ast[i] is None:
                w_ast[i] = 0

        return w_ast, b_ast, eps_ast
    else:
        print('FindSeparator infeasible')
        return None, None, None


def use_sklearn(x, y):
    # 2.1データ整理
    columns_list = x.columns
    X = x[columns_list]
    Y = pd.DataFrame(y)

    # # 2.2学習
    model = LinearRegression()
    model.fit(X, Y)
    w = []
    for i in range(len(columns_list)):
        w.append(model.coef_[0][i])
    b = model.intercept_
    print('coefficient = ', model.coef_[0])  # 説明変数の係数を出力
    print('intercept = ', model.intercept_)  # 切片を出力
    print(w)

    return w, b


def count_s(y):
    s, s_p = 0, 0
    for a in y:
        if a == 0:
            s += 1
        else:
            s_p += 1

    return s, s_p


def sort_dataset(x_df, y, w, b):
    z, z_p = [], []
    # print(x_df)
    # print(y)
    # print(w)
    # TODO: eps<1の時のバグ対処----
    # if len(x_df) == 0:
    #     print(x_df)
    #     return z, z_p
    # -------------------------
    for index, a in y.iteritems():
        if a == 0:
            z.append(naiseki(w, x_df.loc[index]) - b)
        elif a == 1:
            z_p.append(naiseki(w, x_df.loc[index]) - b)
    if y is None:
        return z, z_p
    z.sort()
    z_p.sort(reverse=True)
    # print(z[:5])
    # print(z_p[:5])

    return z, z_p


def naiseki(a, b):
    result = 0
    for i in range(len(a)):
        result += a[i] * b[i]

    return result


def find_r(z, z_p):
    r, r_p = -1, -1
    for i in range(len(z)):
        if z[i] < 0:
            r = i
        else:
            # print('all z is grater than 0')
            break
    for i in range(len(z_p)):
        if z_p[i] > 0:
            r_p = i
        else:
            # print('all z is smaller than 0')
            break

    return r, r_p


def find_index_l(z, r, z_p, r_p, s, s_p):
    rho = RHO
    theta = THETA
    for l in range(r, 0, -1):
        if siki_1(l, z, z_p, r, r_p, s, s_p, rho) == True:
            # print(f"l={l}")
            # print(f"r={r}")
            c_A = -z[l]
            break
        c_A = -z[math.floor(r * theta)]

    for l in range(r_p, 0, -1):
        if siki_2(l, z, z_p, r, r_p, s, s_p, rho) == True:
            # print(f"l={l}")
            # print(f"r\'={r_p}")
            c_B = -z_p[l]  # TODO: type miss???
            break
        c_B = -z_p[math.floor(r_p * theta)]

    if r <= 0:
        c_A = 0
    if r_p <= 0:
        c_B = 0

    return c_A, c_B


def siki_1(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0  # |l|
    for j in range(r_p + 1, s_p):
        if z_p[j] <= z[l]:
            temp += 1
    if temp / l <= rho * s_p / s:
        return True

    return False


def siki_2(l, z, z_p, r, r_p, s, s_p, rho):
    temp = 0
    for j in range(r + 1, s):
        if z[j] >= z_p[l]:
            temp += 1
    if temp / l <= rho * s / s_p:
        return True

    return False


def redefine_func(x_df, y, w, b, c_A, c_B, CIDs, a_score):
    for index, vector_x in x_df.iterrows():
        # ver2. 正解を気にしない
        if naiseki(w, vector_x) - b <= -c_A:
            x_df.drop(index, axis=0, inplace=True)
            y.drop(index, inplace=True)
            a_score[CIDs.loc[index]] = 0
            CIDs.drop(index, inplace=True)

        elif naiseki(w, vector_x) - b >= c_B:
            x_df.drop(index, axis=0, inplace=True)
            y.drop(index, inplace=True)
            a_score[CIDs.loc[index]] = 1
            CIDs.drop(index, inplace=True)

    D = len(y)

    return D, x_df, y


def experiment_test(x_test, y_test, w, b, CIDs_test, a_score_test):
    # 3.1 s, s'を数える
    s, s_p = count_s(y_test)
    # 3.2 ソートする
    z, z_p = sort_dataset(x_test, y_test, w, b)
    # 3.3 r, r'を探す
    r, r_p = find_r(z, z_p)
    # 3.4 index(l)を探す
    c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p)
    # 3.5 振り分け
    for index, vector_x in x_test.iterrows():
        if naiseki(w, vector_x) - b <= -c_A:
            x_test.drop(index, axis=0, inplace=True)
            y_test.drop(index, inplace=True)
            a_score_test[CIDs_test.loc[index]] = 0
            CIDs_test.drop(index, inplace=True)

        elif naiseki(w, vector_x) - b >= c_B:
            x_test.drop(index, axis=0, inplace=True)
            y_test.drop(index, inplace=True)
            a_score_test[CIDs_test.loc[index]] = 1
            CIDs_test.drop(index, inplace=True)

    new_D = len(x_test)
    return new_D, x_test, y_test


def calc_a_q(D, K, a, x):
    a_q = [0] * K
    for d in range(K):
        temp_A = 0
        temp_B = 0
        # for i in range(D):
        #     if a[i] == 0 and x[i][d] isinD_q:
        #         temp_A += 1
        #     elif a[i] == 1 and x[i][d] isinD_q:
        #         temp_B += 1
        if temp_A > temp_B:
            a_q[d] = 0
        elif temp_A < temp_B:
            a_q[d] = 1

    return a_q


def constructing_DT_based_HP(x_df, y, D, K, w_p, b_p, c_p_A, c_p_B, CIDs, a_score):
    # 2. 超平面（hyper plane）を探す
    # w, b = use_sklearn(x, y)
    w, b, eps = find_separator(x_df, y, D, K, w_p, b_p, CIDs)
    # print(f"w={w}")
    # print(f"b={b}")
    print(f"eps={eps}")

    # 3. 決定木の実装
    # 3.1 s, s'を数える
    s, s_p = count_s(y)

    # 3.2 ソートする
    z, z_p = sort_dataset(x_df, y, w, b)
    # print(s, len(z))
    # print(s_p, len(z_p))

    # 3.3 r, r'を探す
    r, r_p = find_r(z, z_p)
    # print(r, r_p)

    # 3.4 index(l)を探す
    c_A, c_B = find_index_l(z, r, z_p, r_p, s, s_p)
    c_p_A.append(c_A)
    c_p_B.append(c_B)
    # print(c_A, c_B)

    # 3.5 関数Φとデータセットの再定義
    D, x_df, y = redefine_func(x_df, y, w, b, c_A, c_B, CIDs, a_score)

    return D, x_df, y


def set_a_q(x_df, y, CIDs, a_score):
    countA, countB = 0, 0
    countB = y.sum()
    countA = len(y) - countB
    if countA >= countB:
        value = 0
    else:
        value = 1

    for index, vector_x in x_df.iterrows():
        a_score[CIDs.loc[index]] = value

    return


depths = []


def main_test(train_data: object, train_value: object, test_data: object, test_value: object) -> object:
    # 1. read data set

    x, y = read_dataset(train_data, train_value)
    x_df = x.reset_index(drop=True)
    y = pd.Series(y)
    y_true_train = copy.deepcopy(y)
    CIDs = pd.Series(list(x.index))
    a_score = pd.Series([-1] * len(CIDs), index=list(CIDs))

    # 2. construct a decision tree using hyper plane
    # マクロ変数
    D = len(x_df)  # データの数
    K = len(x_df.columns)  # ベクトルサイズ（記述示の数）
    N_least = N_LEAST  # 許容
    p = 0

    w_p = []
    b_p = []
    c_p_A = []
    c_p_B = []

    while D > N_least:
        p += 1
        print(f"|D|={D}", f"p={p}")
        new_D, new_x_df, new_y = constructing_DT_based_HP(x_df, y, D, K, w_p, b_p, c_p_A, c_p_B, CIDs, a_score)
        D = new_D
        x_df = new_x_df.reset_index(drop=True)
        y = new_y.reset_index(drop=True)
        CIDs.reset_index(drop=True, inplace=True)
    q = p + 1

    # TODO: a_qの処理
    set_a_q(x_df, y, CIDs, a_score)

    a_score_train = a_score.to_numpy()
    y_train = y_true_train.to_numpy()
    train_score = roc_auc_score(y_train, a_score_train)
    bacc_train_score = balanced_accuracy_score(y_train, a_score_train)

    # 3. test ---------------------------------------------
    x_test, y_test = read_dataset(test_data, test_value)
    x_df = x_test.reset_index(drop=True)
    y = pd.Series(y_test)
    CIDs = pd.Series(list(x_df.index))
    a_score_test = pd.Series([-1] * len(CIDs), index=list(CIDs))

    D = len(x_df)
    x_test = x_df.reset_index(drop=True)
    y_true_test = copy.deepcopy(y)
    y_test = y.reset_index(drop=True)
    CIDs.reset_index(drop=True, inplace=True)

    print(len(b_p))
    depths.append(len(b_p))
    for p in range(len(b_p)):
        print(D)
        new_D, new_x_df, new_y = experiment_test(x_test, y_test, w_p[p], b_p[p], CIDs, a_score_test)
        D = new_D
        x_test = new_x_df.reset_index(drop=True)
        y_test = new_y.reset_index(drop=True)
        CIDs.reset_index(drop=True, inplace=True)

    set_a_q(x_test, y_test, CIDs, a_score_test)

    # 4. 結果
    a_score_test = a_score_test.to_numpy()
    y_true_test = y_true_test.to_numpy()
    test_score = roc_auc_score(y_true_test, a_score_test)
    bacc_test_score = balanced_accuracy_score(y_true_test, a_score_test)

    # print("======================================================")
    # print(train_data)
    print(f"train score : {train_score}")
    print(f"test score : {test_score}")

    return train_score, test_score, bacc_train_score, bacc_test_score


def output_xlx(ws, i, data_name, ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score):
    ws["A" + str(i + 2)] = data_name
    ws["B" + str(i + 2)] = ROCAUC_train_score
    ws["C" + str(i + 2)] = ROCAUC_test_score
    ws["D" + str(i + 2)] = BACC_train_score
    ws["E" + str(i + 2)] = BACC_test_score

    return


def wright_columns(ws):
    ws["B1"] = "train score ROC/AOC"
    ws["C1"] = "test score ROC/AOC"
    ws["D1"] = "train score BACC"
    ws["E1"] = "test score BACC"
    return


def wright_parameter(ws):
    ws["G2"] = f"rho = {RHO}"
    ws["G3"] = f"theta = {THETA}"
    ws["G4"] = f"n_least = {N_LEAST}"
    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/TEST/" + y_m_d)

    if not p_file.exists():
        p_file.mkdir()

    return y_m_d


def prepare_output_file():
    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')
    print(date_time)
    file_name = f"outputfile/TEST/{y_m_d}/{date_time}.xlsx"

    return file_name


def check_exist_dataset(fail1, fail2, fail3, fail4):
    if os.path.exists(fail1) == False or os.path.exists(fail2) == False or os.path.exists(
            fail3) == False or os.path.exists(fail4) == False:
        return False

    return True


def main(rho_arg, theta_arg):
    global RHO
    global THETA
    RHO, THETA = rho_arg, theta_arg
    # エクセルシートを用意
    wbname = prepare_output_file()
    wb = excel.Workbook()
    ws = wb.active
    wright_columns(ws)
    wright_parameter(ws)

    # 1.　全てのデータを実験--------------------------------------------------
    TRAIN_CSV, TRAIN_TXT, TEST_CSV, TEST_TXT = read_data_list()
    results = []
    # for i in range(len(TRAIN_CSV)):
    for i in range(1):
        if not check_exist_dataset(TRAIN_CSV[i], TRAIN_TXT[i], TEST_CSV[i], TEST_TXT[i]):
            continue
        train_score1, test_score1, bacc_train_score1, bacc_test_score1 = main_test(train_data=TRAIN_CSV[i],
                                                                                   train_value=TRAIN_TXT[i],
                                                                                   test_data=TEST_CSV[i],
                                                                                   test_value=TEST_TXT[i])
        train_score2, test_score2, bacc_train_score2, bacc_test_score2 = main_test(train_data=TEST_CSV[i],
                                                                                   train_value=TEST_TXT[i],
                                                                                   test_data=TRAIN_CSV[i],
                                                                                   test_value=TRAIN_TXT[i])

        train_score = statistics.median([train_score1, train_score2])
        test_score = statistics.median([test_score1, test_score2])
        bacc_train_score = statistics.median([bacc_train_score1, bacc_train_score2])
        bacc_test_score = statistics.median([bacc_test_score1, bacc_test_score2])
        print("---------------------------------------")
        print(TRAIN_CSV[i])
        print(f"train score median : {train_score}")
        print(f"test score median : {test_score}")
        results.append(test_score)
        output_xlx(ws, i, TRAIN_CSV[i], train_score, test_score, bacc_train_score, bacc_test_score)

    # print(f"average score : {statistics.median(results)}")

    # 2. 一つだけ実験---------------------------------------------------------
    # TRAIN_CSV, TRAIN_TXT, TEST_CSV, TEST_TXT = INPUT_DATA, VALUE_DATA, TEST_INPUT_DATA, TEST_VALUE_DATA
    # train_score1, test_score1, bacc_train_score1, bacc_test_score1 = main_test(train_data=TRAIN_CSV, train_value=TRAIN_TXT, test_data=TEST_CSV, test_value=TEST_TXT)
    # train_score2, test_score2, bacc_train_score2, bacc_test_score2 = main_test(train_data=TEST_CSV, train_value=TEST_TXT, test_data=TRAIN_CSV, test_value=TRAIN_TXT)
    #
    # train_score = statistics.median([train_score1, train_score2])
    # test_score = statistics.median([test_score1, test_score2])
    # bacc_train_score = statistics.median([bacc_train_score1, bacc_train_score2])
    # bacc_test_score = statistics.median([bacc_test_score1, bacc_test_score2])
    # print("---------------------------------------")
    # print(TRAIN_CSV)
    # print(f"train score median : {train_score}")
    # print(f"test score median : {test_score}")
    # output_xlx(ws, 1, TRAIN_CSV, train_score, test_score, bacc_train_score, bacc_test_score)
    # --------------------------------------------------------------------

    wb.save(wbname)

    return train_score, test_score, bacc_train_score, bacc_test_score


if __name__ == "__main__":
    main(RHO, THETA)
