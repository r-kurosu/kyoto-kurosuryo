import time

import numpy as np
import matplotlib.pyplot as plt

import read_datasets, dt_for_cv
import pathlib
import openpyxl as excel
import datetime
import time
import random
import os
import sys
import io

# sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
# sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

rho_list = [0, 0.01, 0.03, 0.05, 0.07, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.5, 0.7, 1]
# theta_list = [0, 0.1, 0.3, 0.5, 0.7, 1]
# lambda_list = [1, 2, 3, 4, 5, 6]
# C_list = [1, 10, 100, 1000, 10000, 100000]
# rho_list = [0, 0.05, 0.5, 1]
# theta_list = [0.1, 1]
# lambda_list = [1, 2]

# rho_list = [0]
theta_list = [0]

lambda_list = [1]
C_list = [1]

# C_list = random.sample(range(0, 10**5, 10), k=1000)
# print(C_list)

CV_TIMES = 1



def wright_columns(ws, dataset):
    gap = 10
    ws.cell(row=1, column=2).value = "test"
    ws.cell(row=1, column=2 + gap).value = "train"

    # 1
    ws.cell(row=2, column=1).value = "rho\\theta"
    ws.cell(row=2, column=1 + gap).value = "rho\\theta"
    for i in range(len(theta_list)):
        ws.cell(row=2, column=i + 2).value = theta_list[i]
        ws.cell(row=2, column=i + 2 + gap).value = theta_list[i]
    for i in range(len(rho_list)):
        ws.cell(row=i + 3, column=1).value = rho_list[i]
        ws.cell(row=i + 3, column=1 + gap).value = rho_list[i]
    ws.cell(row=3 + len(lambda_list), column=2).value = "(lambda=1)"

    # 2
    ws.cell(row=2 + gap, column=1).value = "rh\\lambda"
    ws.cell(row=2 + gap, column=1 + gap).value = "rho\\lambda"
    for i in range(len(lambda_list)):
        ws.cell(row=2 + gap, column=i + 2).value = lambda_list[i]
        ws.cell(row=2 + gap, column=i + 2 + gap).value = lambda_list[i]
    for i in range(len(rho_list)):
        ws.cell(row=i + 3 + gap, column=1).value = rho_list[i]
        ws.cell(row=i + 3 + gap, column=1 + gap).value = rho_list[i]
    ws.cell(row=3 + gap + len(rho_list), column=2).value = "(rho=0.05)"

    # 3
    ws.cell(row=2 + gap*2, column=1).value = "lambda\\theta"
    ws.cell(row=2 + gap*2, column=1 + gap).value = "lambda\\theta"
    for i in range(len(theta_list)):
        ws.cell(row=2 + gap*2, column=i + 2).value = theta_list[i]
        ws.cell(row=2 + gap*2, column=i + 2 + gap).value = theta_list[i]
    for i in range(len(lambda_list)):
        ws.cell(row=i + 3 + gap*2, column=1).value = lambda_list[i]
        ws.cell(row=i + 3 + gap*2, column=1 + gap).value = lambda_list[i]
    ws.cell(row=3 + gap*2 + len(lambda_list), column=2).value = "(theta=0.1)"

    # ws.cell(row=12, column=2).value = dataset

    return


def make_dir(now_time):
    y_m_d = now_time.strftime('%Y-%m-%d')
    p_file = pathlib.Path("outputfile/CV/" + y_m_d)
    p_file_ht = pathlib.Path("outputfile/CV/" + y_m_d + "/hyper_turning")
    p_file_ht_memo = pathlib.Path("outputfile/CV/" + y_m_d + "/ht_memo")

    if not p_file.exists():
        p_file.mkdir()
    if not p_file_ht.exists():
        p_file_ht.mkdir()
    if not p_file_ht_memo.exists():
        p_file_ht_memo.mkdir()
    return y_m_d


def prepare_output_file(data_name):
    target_name = data_name.split(sep="/")

    # 出力用のファイルを準備
    now_time = datetime.datetime.now()
    y_m_d = make_dir(now_time)
    date_time = now_time.strftime('%Y%m%d-%H%M%S')

    file_name = f"outputfile/CV/{y_m_d}/hyper_turning/{date_time}_ht_cv_{target_name[2]}.xlsx"

    return file_name


def edit_data_name(data_name):
    name_list = data_name.split("/")
    property_list = name_list[2].split("_")
    name = property_list[0]

    return name


def plot_func_for_rho_list(train_scores, test_scores, data_name):
    name = edit_data_name(data_name)

    plt.scatter(rho_list, train_scores, color="blue", label="train")
    plt.scatter(rho_list, test_scores, color="orange", label="test")
    plt.ylim(0, 1)
    plt.title(f"BACC score ({name})")
    plt.xlabel(f'rho={rho_list}')
    plt.ylabel('BACC score')
    plt.legend()
    plt.grid()

    plt.savefig(f"score_{name}.png")
    # plt.show()

    return


def main():
    INPUT_CSV, INPUT_TXT = read_datasets.read_data_list_for_cv()
    wb_name = [0]*len(INPUT_CSV)
    gap = 10

    for k, data in enumerate(INPUT_CSV):
        ht_start_time = time.time()

        # エクセルシートを用意
        wb_name[k] = prepare_output_file(data_name=data)
        wb = excel.Workbook()
        ws = wb.active
        ws.title = "BACC score"
        wright_columns(ws, data)
        wb.create_sheet(title="AUC score")
        ws2 = wb["AUC score"]
        wright_columns(ws2, data)

        bacc_score_data_test = np.zeros((len(rho_list), len(theta_list), len(lambda_list), len(C_list)))
        bacc_score_data_train = np.zeros((len(rho_list), len(theta_list), len(lambda_list), len(C_list)))
        auc_score_data_test = np.zeros((len(rho_list), len(theta_list), len(lambda_list), len(C_list)))
        auc_score_data_train = np.zeros((len(rho_list), len(theta_list), len(lambda_list), len(C_list)))

        train_scores = []
        test_scores = []

        # --
        for i, rho in enumerate(rho_list):
            for j, theta in enumerate(theta_list):
                for l, lambd in enumerate(lambda_list):
                    for m, C in enumerate(C_list):
                        st_time = time.time()
                        print(f"rho:{rho},theta:{theta}, lambda:{lambd}, C:{C}")
                        ROCAUC_train_score, ROCAUC_test_score, BACC_train_score, BACC_test_score, max_depth \
                            = dt_for_cv.main(rho, theta, lambd, C, INPUT_CSV[k], INPUT_TXT[k], CV_TIMES)

                        # 出力
                        if lambd == 1:
                            ws.cell(row=i + 3, column=j + 2).value = BACC_test_score
                            ws.cell(row=i + 3, column=j + 2 + gap).value = BACC_train_score
                            ws2.cell(row=i + 3, column=j + 2).value = ROCAUC_test_score
                            ws2.cell(row=i + 3, column=j + 2 + gap).value = ROCAUC_train_score

                        # if rho == 0.05:
                        #     ws.cell(row=j + 3 + gap, column=l + 2).value = BACC_test_score
                        #     ws.cell(row=j + 3 + gap, column=l + 2 + gap).value = BACC_train_score
                        #     ws2.cell(row=j + 3 + gap, column=l + 2).value = ROCAUC_test_score
                        #     ws2.cell(row=j + 3 + gap, column=l + 2 + gap).value = ROCAUC_train_score
                        #
                        # if theta == 0.1:
                        #     ws.cell(row=l + 3 + gap*2, column=i + 2).value = BACC_test_score
                        #     ws.cell(row=l + 3 + gap*2, column=i + 2 + gap).value = BACC_train_score
                        #     ws2.cell(row=l + 3 + gap*2, column=i + 2).value = ROCAUC_test_score
                        #     ws2.cell(row=l + 3 + gap*2, column=i + 2 + gap).value = ROCAUC_train_score

                        bacc_score_data_test[i, j, l, m] = BACC_test_score
                        bacc_score_data_train[i, j, l, m] = BACC_train_score
                        auc_score_data_test[i, j, l, m] = ROCAUC_test_score
                        auc_score_data_train[i, j, l, m] = ROCAUC_train_score

                        train_scores.append(BACC_train_score)
                        test_scores.append(BACC_test_score)

                        ed_time = time.time()
                        print("time {:.1f}".format(ed_time - st_time))

        wb.save(wb_name[k])

        bacc_max_test_index = np.unravel_index(np.argmax(bacc_score_data_test), bacc_score_data_test.shape)
        bacc_max_train_index = np.unravel_index(np.argmax(bacc_score_data_train), bacc_score_data_train.shape)
        auc_max_train_index = np.unravel_index(np.argmax(auc_score_data_train), auc_score_data_train.shape)
        auc_max_test_index = np.unravel_index(np.argmax(auc_score_data_test), auc_score_data_test.shape)
        
        # print(max_test_index, max_train_index)
        bacc_max_test_score = bacc_score_data_test[bacc_max_test_index]
        bacc_max_train_score = bacc_score_data_train[bacc_max_train_index]
        auc_max_test_score = auc_score_data_test[auc_max_test_index]
        auc_max_train_score = auc_score_data_train[auc_max_train_index]
        # print(max_test_score, max_train_score)
        # print(score_data_train)
        # print(score_data_test)

        ht_end_time = time.time()

        print("*"*70)
        print(f"{data}")
        print(f"max train score (BACC) = {bacc_max_train_score}, (rho={rho_list[bacc_max_train_index[0]]}, theta={theta_list[bacc_max_train_index[1]]})")
        print(f"max test score (BACC) = {bacc_max_test_score}, (rho={rho_list[bacc_max_test_index[0]]}, theta={theta_list[bacc_max_test_index[1]]})")
        print(f"max train score (AUC) = {auc_max_train_score}, (rho={rho_list[auc_max_train_index[0]]}, theta={theta_list[auc_max_train_index[1]]})")
        print(f"max test score (AUC) = {auc_max_test_score}, (rho={rho_list[auc_max_test_index[0]]}, theta={theta_list[auc_max_test_index[1]]})")
        print("計算時間 {:.1f}".format(ht_end_time - ht_start_time))
        print("*"*70)

        plot_func_for_rho_list(train_scores, test_scores, data_name=data)

    return


if __name__ == "__main__":
    main()
